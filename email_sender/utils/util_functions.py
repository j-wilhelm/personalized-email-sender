import openpyxl as oxl

def parse_excel(filepath):
    # Has to loop over rows.
    # Append each relevant object to a dictionary
    # dict structury: {emailaddres: {name: "x"}, {"info":

    # Loop over email addresses:
    wb = oxl.load_workbook("test_excel.xlsx")
    current_sheet = wb["Current"]

    colnames_dict = {}
    column_index = 0
    for col in current_sheet.iter_cols(1, current_sheet.max_column):
        colname = col[0].value
        if colname is None:
            column_index += 1
            continue
        colnames_dict[colname] = column_index
        column_index += 1

    email_dict = {}
    row_index = 0
    for row in current_sheet.iter_rows(1, current_sheet.max_rows):
        rowname = row[0].value
        if rowname is None:
            row_index += 1
            continue
        email_dict[rowname] = row_index
        row_index += 1

    member_dictionaries = []
    for email_address, row_ix in email_dict.items():
        member_dict = init_dict(email_address)

        # Loop through other info
        for colname, col_ix in colnames_dict.items():
            if colname == "Email":
                continue

            if colname == "name":
                member_dict["name"] = current_sheet[row_ix][col_ix].value
            elif colname == "totaal":
                member_dict["totaal"] = current_sheet[row_ix][col_ix].value
            else:
                member_dict["info"][colname] = current_sheet[row_ix][col_ix].value

        member_dictionaries.append(member_dict)

def init_dict(emailaddress):
    dict = {}
    dict[emailaddress] = {}
    dict[emailaddress]["name"] = None
    dict[emailaddress]["info"] = dict
    dict[emailaddress]["totaal"] = 0